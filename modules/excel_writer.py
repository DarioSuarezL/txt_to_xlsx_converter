import re
from openpyxl.styles import PatternFill
from modules.excel_formatter import colorear_fila

def procesar_datos(lineas, ws, wb):
    """
    Procesa las líneas de un archivo de texto.
    
    Args:
        lineas (list): Lista de líneas del archivo

    Returns:
        pd.DataFrame: DataFrame con los datos procesados
    """

    ws['A1'] = "REGISTRO"
    ws['B1'] = "ALUMNO"
    ws['C1'] = "SIGLA"
    ws['D1'] = "CARRERA"
    ws['E1'] = "Ingreso"
    ws['F1'] = "PPAC"
    ws['G1'] = "PPACE"
    ws['H1'] = "S/Ano"
    ws['I1'] = "PPS"
    ws['J1'] = "MATERIA 1"
    ws['K1'] = "NOTA 1"
    ws['L1'] = "MATERIA 2"
    ws['M1'] = "NOTA 2"
    ws['N1'] = "MATERIA 3"
    ws['O1'] = "NOTA 3"
    ws['P1'] = "MATERIA 4"
    ws['Q1'] = "NOTA 4"
    ws['R1'] = "MATERIA 5"
    ws['S1'] = "NOTA 5"
    ws['T1'] = "MATERIA 6"
    ws['U1'] = "NOTA 6"
    ws['V1'] = "MATERIA 7"
    ws['W1'] = "NOTA 7"
    ws['X1'] = "MATERIA 8"
    ws['Y1'] = "NOTA 8"
    ws['Z1'] = "MATERIA 9"
    ws['AA1'] = "NOTA 9"
    ws['AB1'] = "MATERIA 10"
    ws['AC1'] = "NOTA 10"

    counter = 2 #fila que inicia

    nombre_contador = 0

    registro = None
    alumno = None
    sigla = None
    carrera = None
    ingreso = None
    ppac = None
    ppace = None

    lineas_a_omitir = [
        "----------------------------------------------------------------------------------------------------",
        "----------------------------------------------------------------------------------CPD-fapB50",
        ""
    ]

    fill = PatternFill(start_color="FFDDEBF7", end_color="FFDDEBF7", fill_type = "solid")


    for linea in lineas:
        debugger = True

    #LINEAS IGNORADAS
        if linea in lineas_a_omitir:
            continue

        if linea.startswith("U.A.G.R.M.") or linea.startswith("S.I.F.") or linea.startswith("SANTA CRUZ") or linea.startswith("FORMA DE INGRESO") or linea.startswith("Carrera Paralela") or linea.startswith("S/Ano"):
            continue
    #lINEAS IGNORADAS

        if counter % 2 == 0:
            colorear_fila(ws, counter, fill)

        #Extrae datos del estudiante
        match_estudiante = re.search(r'#\s+\d+:\s+Estudiante:\s*(\d+)\s+-\s+(.+)', linea)
        if match_estudiante:
            registro, alumno = match_estudiante.groups()
            nombre_contador += 1
            # print(f"# {nombre_contador}: Estudiante: {alumno}")
            debugger = False
            continue
        else:
            #Caso especial de estudiante que tiene registro corto
            match_estudiante = re.search(r'#\s+\d+:\s+Estudiante:\s+(\d+)\s+-\s+(.+)', linea)
            if match_estudiante:
                registro, alumno = match_estudiante.groups()
                debugger = False
                continue

        if registro and alumno:
            ws['A'+str(counter)] = registro
            ws['B'+str(counter)] = alumno
        
        #Extrae datos de la carrera
        match_carrera = re.search(r'Carrera Actual\s+:\s+(\S+)\s+(.+?)\s+INGRESO:\s+(\S+)\s+PPAC:\s+(\d+)\s+PPACE:\s+(\d+)', linea)
        if match_carrera:
            sigla, carrera, ingreso, ppac, ppace = match_carrera.groups()
            debugger = False
            continue

        if sigla and carrera and ingreso and ppac and ppace:
            # ws['C'+str(counter)] = sigla    #SE OCUPARÁ LA CARRERA DE MATERIA Y NO LA DE ESTUDIANTE
            ws['D'+str(counter)] = carrera
            ws['E'+str(counter)] = ingreso
            ws['F'+str(counter)] = ppac
            ws['G'+str(counter)] = ppace


        #Extrae datos del histórico
        #Si tenia 10 inscritas
        match_historico = re.search(r"^([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)", linea)
        if match_historico:
            materia8, nota8, materia9, nota9, materia10, nota10 = match_historico.groups()
            ws['X'+str(counter-1)] = materia8
            ws['Y'+str(counter-1)] = nota8
            ws['Z'+str(counter-1)] = materia9
            ws['AA'+str(counter-1)] = nota9
            ws['AB'+str(counter-1)] = materia10
            ws['AC'+str(counter-1)] = nota10
            debugger = False
            continue

        #Si tenia 9 inscritas
        match_historico = re.search(r"^([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)", linea)
        if match_historico:
            materia8, nota8, materia9, nota9 = match_historico.groups()
            ws['X'+str(counter-1)] = materia8
            ws['Y'+str(counter-1)] = nota8
            ws['Z'+str(counter-1)] = materia9
            ws['AA'+str(counter-1)] = nota9
            debugger = False
            continue

        #Si tenia 8 inscritas
        match_historico = re.search(r"^([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)", linea)
        if match_historico:
            materia8, nota8 = match_historico.groups()
            ws['X'+str(counter-1)] = materia8
            ws['Y'+str(counter-1)] = nota8
            debugger = False
            continue

        #Si tenia 7 inscritas
        match_historico = re.search(r"([A-Za-z0-9-@#%]/\d{4})\s*(\d+-[A-Za-z0-9-@#%])\s*\(\s*(\d+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)", linea)
        if match_historico:
            s_ano, sigla_carrera, pps, materia1, nota1, materia2, nota2, materia3, nota3, materia4, nota4, materia5, nota5, materia6, nota6, materia7, nota7 = match_historico.groups()
            ws['C'+str(counter)] = sigla_carrera

            ws['H'+str(counter)] = s_ano
            ws['I'+str(counter)] = pps
            ws['J'+str(counter)] = materia1
            ws['K'+str(counter)] = nota1
            ws['L'+str(counter)] = materia2
            ws['M'+str(counter)] = nota2
            ws['N'+str(counter)] = materia3
            ws['O'+str(counter)] = nota3
            ws['P'+str(counter)] = materia4
            ws['Q'+str(counter)] = nota4
            ws['R'+str(counter)] = materia5
            ws['S'+str(counter)] = nota5
            ws['T'+str(counter)] = materia6
            ws['U'+str(counter)] = nota6
            ws['V'+str(counter)] = materia7
            ws['W'+str(counter)] = nota7

            counter += 1
            debugger = False
            continue
        
        #Si tenia 6 inscritas
        match_historico = re.search(r"([A-Za-z0-9-@#%]/\d{4})\s*(\d+-[A-Za-z0-9-@#%])\s*\(\s*(\d+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)", linea)
        if match_historico:
            s_ano, sigla_carrera, pps, materia1, nota1, materia2, nota2, materia3, nota3, materia4, nota4, materia5, nota5, materia6, nota6 = match_historico.groups()
            ws['C'+str(counter)] = sigla_carrera
            
            ws['H'+str(counter)] = s_ano
            ws['I'+str(counter)] = pps
            ws['J'+str(counter)] = materia1
            ws['K'+str(counter)] = nota1
            ws['L'+str(counter)] = materia2
            ws['M'+str(counter)] = nota2
            ws['N'+str(counter)] = materia3
            ws['O'+str(counter)] = nota3
            ws['P'+str(counter)] = materia4
            ws['Q'+str(counter)] = nota4
            ws['R'+str(counter)] = materia5
            ws['S'+str(counter)] = nota5
            ws['T'+str(counter)] = materia6
            ws['U'+str(counter)] = nota6

            counter += 1
            debugger = False
            continue
        
        #Si tenia 5 inscritas
        match_historico = re.search(r"([A-Za-z0-9-@#%]/\d{4})\s*(\d+-[A-Za-z0-9-@#%])\s*\(\s*(\d+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)", linea)
        if match_historico:
            s_ano, sigla_carrera, pps, materia1, nota1, materia2, nota2, materia3, nota3, materia4, nota4, materia5, nota5 = match_historico.groups()
            ws['C'+str(counter)] = sigla_carrera
            
            ws['H'+str(counter)] = s_ano
            ws['I'+str(counter)] = pps
            ws['J'+str(counter)] = materia1
            ws['K'+str(counter)] = nota1
            ws['L'+str(counter)] = materia2
            ws['M'+str(counter)] = nota2
            ws['N'+str(counter)] = materia3
            ws['O'+str(counter)] = nota3
            ws['P'+str(counter)] = materia4
            ws['Q'+str(counter)] = nota4
            ws['R'+str(counter)] = materia5
            ws['S'+str(counter)] = nota5

            counter += 1
            debugger = False
            continue

        #Si tenia 4 inscritas
        match_historico = re.search(r"([A-Za-z0-9-@#%]/\d{4})\s*(\d+-[A-Za-z0-9-@#%])\s*\(\s*(\d+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)", linea)
        if match_historico:
            s_ano, sigla_carrera, pps, materia1, nota1, materia2, nota2, materia3, nota3, materia4, nota4 = match_historico.groups()
            ws['C'+str(counter)] = sigla_carrera
            
            ws['H'+str(counter)] = s_ano
            ws['I'+str(counter)] = pps
            ws['J'+str(counter)] = materia1
            ws['K'+str(counter)] = nota1
            ws['L'+str(counter)] = materia2
            ws['M'+str(counter)] = nota2
            ws['N'+str(counter)] = materia3
            ws['O'+str(counter)] = nota3
            ws['P'+str(counter)] = materia4
            ws['Q'+str(counter)] = nota4

            counter += 1
            debugger = False
            continue

        #Si tenia 3 inscritas
        match_historico = re.search(r"([A-Za-z0-9-@#%]/\d{4})\s*(\d+-[A-Za-z0-9-@#%])\s*\(\s*(\d+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)", linea)
        if match_historico:
            s_ano, sigla_carrera, pps, materia1, nota1, materia2, nota2, materia3, nota3 = match_historico.groups()
            ws['C'+str(counter)] = sigla_carrera
            
            ws['H'+str(counter)] = s_ano
            ws['I'+str(counter)] = pps
            ws['J'+str(counter)] = materia1
            ws['K'+str(counter)] = nota1
            ws['L'+str(counter)] = materia2
            ws['M'+str(counter)] = nota2
            ws['N'+str(counter)] = materia3
            ws['O'+str(counter)] = nota3

            counter += 1
            debugger = False
            continue

        #Si tenia 2 inscritas
        match_historico = re.search(r"([A-Za-z0-9-@#%]/\d{4})\s*(\d+-[A-Za-z0-9-@#%])\s*\(\s*(\d+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)", linea)
        if match_historico:
            s_ano, sigla_carrera, pps, materia1, nota1, materia2, nota2 = match_historico.groups()
            ws['C'+str(counter)] = sigla_carrera
            
            ws['H'+str(counter)] = s_ano
            ws['I'+str(counter)] = pps
            ws['J'+str(counter)] = materia1
            ws['K'+str(counter)] = nota1
            ws['L'+str(counter)] = materia2
            ws['M'+str(counter)] = nota2

            counter += 1
            debugger = False
            continue

        #Si tenia 1 inscritas
        match_historico = re.search(r"([A-Za-z0-9-@#%]/\d{4})\s*(\d+-[A-Za-z0-9-@#%])\s*\(\s*(\d+)\s*\)[a-z]*\s*([A-Za-z0-9-@#%]+)\s*\(\s*([A-Za-z0-9-]+)\s*\)", linea)
        if match_historico:
            s_ano, sigla_carrera, pps, materia1, nota1 = match_historico.groups()
            ws['C'+str(counter)] = sigla_carrera
            
            ws['H'+str(counter)] = s_ano
            ws['I'+str(counter)] = pps
            ws['J'+str(counter)] = materia1
            ws['K'+str(counter)] = nota1

            counter += 1
            debugger = False
            continue

        if debugger:
            print(f"Error en la línea: {linea}")

    print(f"El documento cuenta con {counter - 1} filas hechas")
    print(f"El documento cuenta con {nombre_contador} estudiantes")

