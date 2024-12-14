from openpyxl.utils import get_column_letter 
from openpyxl.styles import PatternFill, Alignment, Font

def colorear_fila(ws, counter, fill):
    ws['A'+str(counter)].fill = fill
    ws['B'+str(counter)].fill = fill
    ws['C'+str(counter)].fill = fill
    ws['D'+str(counter)].fill = fill
    ws['E'+str(counter)].fill = fill
    ws['F'+str(counter)].fill = fill
    ws['G'+str(counter)].fill = fill
    ws['H'+str(counter)].fill = fill
    ws['I'+str(counter)].fill = fill
    ws['J'+str(counter)].fill = fill
    ws['K'+str(counter)].fill = fill
    ws['L'+str(counter)].fill = fill
    ws['M'+str(counter)].fill = fill
    ws['N'+str(counter)].fill = fill
    ws['O'+str(counter)].fill = fill
    ws['P'+str(counter)].fill = fill
    ws['Q'+str(counter)].fill = fill
    ws['R'+str(counter)].fill = fill
    ws['S'+str(counter)].fill = fill
    ws['T'+str(counter)].fill = fill
    ws['U'+str(counter)].fill = fill
    ws['V'+str(counter)].fill = fill
    ws['W'+str(counter)].fill = fill
    ws['X'+str(counter)].fill = fill
    ws['Y'+str(counter)].fill = fill
    ws['Z'+str(counter)].fill = fill
    ws['AA'+str(counter)].fill = fill
    ws['AB'+str(counter)].fill = fill
    ws['AC'+str(counter)].fill = fill


def ajustar_tamanio_columnas(ws):
    """
    Ajusta el tamaño de las columnas de un DataFrame en un archivo Excel.
    
    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): Hoja de cálculo de Excel
    """
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 4)
        ws.column_dimensions[column_letter].width = adjusted_width
    print("Tamaño de columnas ajustado correctamente.")


def formatear_celdas(ws):
    """
    Colorea las celdas de un DataFrame en un archivo Excel.
    
    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): Hoja de cálculo de Excel
    """
    fill = PatternFill(start_color="FF5B9BD5", end_color="FF5B9BD5", fill_type = "solid")
    font = Font(color="FFFFFFFF")
    alignment = Alignment(horizontal='center', vertical='center')

    ws['A1'].fill = fill
    ws['B1'].fill = fill
    ws['C1'].fill = fill
    ws['D1'].fill = fill
    ws['E1'].fill = fill
    ws['F1'].fill = fill
    ws['G1'].fill = fill
    ws['H1'].fill = fill
    ws['I1'].fill = fill
    ws['J1'].fill = fill
    ws['K1'].fill = fill
    ws['L1'].fill = fill
    ws['M1'].fill = fill
    ws['N1'].fill = fill
    ws['O1'].fill = fill
    ws['P1'].fill = fill
    ws['Q1'].fill = fill
    ws['R1'].fill = fill
    ws['S1'].fill = fill
    ws['T1'].fill = fill
    ws['U1'].fill = fill
    ws['V1'].fill = fill
    ws['W1'].fill = fill
    ws['X1'].fill = fill
    ws['Y1'].fill = fill
    ws['Z1'].fill = fill
    ws['AA1'].fill = fill
    ws['AB1'].fill = fill
    ws['AC1'].fill = fill

    ws['A1'].font = font
    ws['B1'].font = font
    ws['C1'].font = font
    ws['D1'].font = font
    ws['E1'].font = font
    ws['F1'].font = font
    ws['G1'].font = font
    ws['H1'].font = font
    ws['I1'].font = font
    ws['J1'].font = font
    ws['K1'].font = font
    ws['L1'].font = font
    ws['M1'].font = font
    ws['N1'].font = font
    ws['O1'].font = font
    ws['P1'].font = font
    ws['Q1'].font = font
    ws['R1'].font = font
    ws['S1'].font = font 
    ws['T1'].font = font  
    ws['U1'].font = font 
    ws['V1'].font = font  
    ws['W1'].font = font 
    ws['X1'].font = font  
    ws['Y1'].font = font 
    ws['Z1'].font = font  
    ws['AA1'].font = font 
    ws['AB1'].font = font  
    ws['AC1'].font = font

    print("Celdas header coloreadas correctamente.")

    ws['A1'].alignment = alignment
    ws['B1'].alignment = alignment
    ws['C1'].alignment = alignment
    ws['D1'].alignment = alignment
    ws['E1'].alignment = alignment
    ws['F1'].alignment = alignment
    ws['G1'].alignment = alignment
    ws['H1'].alignment = alignment
    ws['I1'].alignment = alignment
    ws['J1'].alignment = alignment
    ws['K1'].alignment = alignment
    ws['L1'].alignment = alignment
    ws['M1'].alignment = alignment
    ws['N1'].alignment = alignment
    ws['O1'].alignment = alignment
    ws['P1'].alignment = alignment
    ws['Q1'].alignment = alignment
    ws['R1'].alignment = alignment
    ws['S1'].alignment = alignment
    ws['T1'].alignment = alignment
    ws['U1'].alignment = alignment
    ws['V1'].alignment = alignment
    ws['W1'].alignment = alignment
    ws['X1'].alignment = alignment
    ws['Y1'].alignment = alignment
    ws['Z1'].alignment = alignment
    ws['AA1'].alignment = alignment
    ws['AB1'].alignment = alignment
    ws['AC1'].alignment = alignment

    print("Celdas header alineadas correctamente.")